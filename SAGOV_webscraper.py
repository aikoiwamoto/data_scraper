import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from openpyxl import load_workbook

# Initialize ChromeDriver
chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome()

# Open Excel file
excel_file = 'INFS5135 - SAGOV_Data_Sheet.xlsx'
sheet_name = 'Data Sheet'
wb = load_workbook(excel_file)
ws = wb[sheet_name]

# Change Excel's calculation mode to automatic
wb.calculation.calcMode = "auto"

cell_row = 1

# Iterate through the values in column B
for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):

    # Open website in Google Chrome
    driver.get('https://account.ezyreg.sa.gov.au/account/check-registration.htm')

    for cell_value in row:
        print(cell_value)

        # Locate the search field on the website and input data with retry
        search_field = None
        max_retries = 3
        for retry in range(max_retries):
            try:
                search_field = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[3]/div[5]/form/div[2]/div/input')))
                search_field.send_keys(cell_value)
                break  # If successful, break out of the loop
            except StaleElementReferenceException:
                if retry < max_retries - 1:
                    # If a StaleElementReferenceException occurs, retry
                    continue
                else:
                    raise  # If all retries fail, raise the exception

        # Submit the search form
        search_field.submit()

        # Wait for the web page to load
        time.sleep(5)

        cell_row += 1
        col_letters = ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
        web_elements = [
            'expiry_date', 'plate_num', 'make', 'body_type', 'prim_colour', 'ctp_insurer', 'ctp_ins_prem_class', 'vin'
        ]

        # Check if certain elements (e.g., the expiry_date element) are found on the page
        elements_found = all(driver.find_elements(By.XPATH, f'/html/body/div[2]/div[3]/div[7]/div[2]/div[2]/div/div[{i}]/div[2]/div') for i in range(1, 7))

        if elements_found:
            
            for i, web_element in enumerate(web_elements):
                try:
                    element_value = driver.find_element(By.XPATH, f'/html/body/div[2]/div[3]/div[7]/div[2]/div[2]/div/div[{i+1}]/div[2]/div').text
                except:
                    element_value = ''

                cell_col = ws[f'{col_letters[i]}{cell_row}']
                cell_col.value = element_value

            next_button = driver.find_element(By.ID, 'step_3_submit')
            next_button.click()

        if not elements_found:
            back_button = driver.find_element(By.ID, 'step-2-1-submit')

            for col_letter in col_letters:
                cell_col = ws[f'{col_letter}{cell_row}']
                cell_col.value = ""

            # Check if the "back" button is present and click it
            if back_button.is_displayed():
                back_button.click()
                cell_r = ws[f'R{cell_row}']
                cell_r.value = f"Multiple cars found for {cell_value}. Skipping to the next value."
            else:
                cell_r = ws[f'R{cell_row}']
                cell_r.value = f"{cell_value} not found. Skipping to the next value."

        # Add timestamp to each Excel row
        timestamp_col = f'S{cell_row}'
        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        cell_m = ws[timestamp_col]
        cell_m.value = timestamp

        wb.save(excel_file)

driver.quit()

print("Web scraping complete.")
