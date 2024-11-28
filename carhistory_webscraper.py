import time
import undetected_chromedriver as uc
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException
from openpyxl import load_workbook

# Initialize ChromeDriver
chrome_options = uc.ChromeOptions()
# chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
# chrome_options.add_argument('--ignore-certificate-errors')
# chrome_options.add_argument()
chrome_options.headless = False

# Create Undetected Chromedriver with options
driver = uc.Chrome(options=chrome_options)

        
# # #Initialize Microsoft Edge WebDriver
# edge_driver_path = "C:\\Users\\aikoi\\Downloads\\msedgedriver.exe"
# driver = webdriver.Edge()

# Open Excel file
excel_file = 'INFS5135 - VIN_Data_Sheet.xlsx'
sheet_name = 'Data Sheet'
wb = load_workbook(excel_file)
ws = wb[sheet_name]

# Change Excel's calculation mode to automatic
wb.calculation.calcMode = "auto"

cell_row = 1

# Iterate through the values in column B
for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):

    for cell_value in row:
        print(cell_value)


        # Open website
        driver.get('https://www.carhistory.com.au/')

        time.sleep(5)

        # Locate the search field on the website and input data with retry
        rego_field = None
        max_retries = 10
        for retry in range(max_retries):
            try:
                rego_field = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, '/html/body/section[1]/div[1]/div/div/form/div/section/form/div[2]/div[2]/div[1]/input'))
                    )
                rego_field.send_keys(cell_value)

                state_field = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, '/html/body/section[1]/div[1]/div/div/form/div/section/form/div[2]/div[2]/div[2]/select'))
                    )
                dropdown = Select(state_field)
                sa_option = 'SA'
                dropdown.select_by_visible_text(sa_option)

                break  # If successful, break out of the loop
            except StaleElementReferenceException:
                if retry < max_retries - 1:
                    # If a StaleElementReferenceException occurs, retry
                    continue
                else:
                    raise  # If all retries fail, raise the exception

        # Submit the search form
        search_button = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/section[1]/div[1]/div/div/form/div/section/form/div[2]/div[2]/div[3]/button'))
            )
        search_button.click()

        # Wait for the web page to load
        time.sleep(15)

        cell_row += 1

        vin_value = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div/div[1]/form/input[2]').get_attribute('value')
        make_value = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div/div[1]/form/div[1]/div/div/div[3]/span').text
        plate_num_state_value = driver.find_element(By.XPATH, '/html/body/div[1]/div[1]/div/div[1]/form/input[8]').get_attribute('value')

        vin_col = ws[f'O{cell_row}']
        make_col = ws[f'P{cell_row}']
        plate_num_state_col = ws[f'R{cell_row}']
        
        vin_col.value = vin_value
        make_col.value = make_value
        plate_num_state_col.value = plate_num_state_value

        # Add timestamp to each Excel row
        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        timestamp_col = ws[f'Z{cell_row}']
        timestamp_col.value = timestamp

        wb.save(excel_file)

driver.quit()

print("Web scraping complete.")
