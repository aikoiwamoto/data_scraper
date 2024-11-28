import time
import re
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException
from openpyxl import load_workbook

# Initialize ChromeDriver
chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
chrome_options.add_argument('--ignore-certificate-errors')
driver = webdriver.Chrome(chrome_options)

# Open Excel file
excel_file = 'INFS5135 - Car_Info_Data_Sheet.xlsx'
sheet_name = 'Data Sheet'
wb = load_workbook(excel_file)
ws = wb[sheet_name]

# Change Excel's calculation mode to automatic
wb.calculation.calcMode = "auto"

# Define a regular expression patterns for the "Variant" and "Engine" data
variant_pattern = r'\d+(?:sp|sped) (Auto|Man) [\d\w\s&]+ [\d\w\s]+ \[.*\]'


cell_row = 1

# Open website in Google Chrome
driver.get('https://www.repco.com.au/')

# time.sleep(5)
# WebDriverWait(driver, 20).until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"//iframe[@title='Widget containing a Cloudflare security challenge']")))
# WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//label[@class='ctp-checkbox-label']"))).click()

# Iterate through the values in column B
for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):

    for cell_value in row:
        print(cell_value)

        time.sleep(5)

        # Locate the search field on the website and input data with retry
        max_retries = 10
        for retry in range(max_retries):
            try:
                set_my_vh_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/main/header/nav[2]/div/div[1]/div[2]/ul/li[4]/div/div[1]/a'))
                    )
                set_my_vh_button.click()

                break  # If successful, break out of the loop
            except StaleElementReferenceException:
                if retry < max_retries - 1:
                    # If a StaleElementReferenceException occurs, retry
                    continue
                else:
                    raise  # If all retries fail, raise the exception


        rego_search_field = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/main/header/nav[2]/div/div[1]/div[2]/ul/li[4]/div/div[3]/div/div[5]/div[4]/div[2]/div[3]/div[1]/input'))
            )
        rego_search_field.send_keys(cell_value)

        state_field = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/main/header/nav[2]/div/div[1]/div[2]/ul/li[4]/div/div[3]/div/div[5]/div[4]/div[2]/div[4]/div[1]/select'))
            )
        dropdown = Select(state_field)
        sa_option = 'South Australia'
        dropdown.select_by_visible_text(sa_option)


        # Submit the search form
        search_button = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/main/header/nav[2]/div/div[1]/div[2]/ul/li[4]/div/div[3]/div/div[5]/div[4]/div[2]/div[5]/div'))
            )
        search_button.click()
  

        cell_row += 1

        # Check if certain elements (e.g., the expiry_date element) are found on the page
        try:
            element = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/main/header/nav[2]/div/div[1]/div[2]/ul/li[4]/div/div[3]/div/div[5]/div[4]/div[2]/div[7]/ul/li/div[1]/span'))
                ).text

            if element:
                car_info_col = ws[f'N{cell_row}']
                car_info_col.value = element

                #Extract and remove "Production Period"
                prod_per_match = re.search(r'\[(.*?)\]', element)
                if prod_per_match:
                    period = prod_per_match.group(1)
                    # Assign the extracted "Production Period" data to the respective column
                    prod_per_col = ws[f'S{cell_row}']
                    prod_per_col.value = period

                    # Remove "Production Period" data from car description element
                    element = re.sub(r'\[(.*?)\]', '', element)
                    element = element.rstrip()

                # Extract and remove "Variant" data
                # variant_match = re.search(r'\d+(?:sp|sped) (Auto|Man) [\d\w\s&]+ [\d\w\s]+ \[.*\]', element)
                variant_match = re.search(r'\d+[a-zA-Z]* (Auto|Man) [\d\w\s&]+ [\d\w\s]+', element)
                if variant_match:
                    variant = variant_match.group(0)

                    # Assign the extracted "Variant" data to the respective column
                    variant_col = ws[f'R{cell_row}']
                    variant_col.value = variant

                    # Remove "Variant" data from car description element
                    element = element.replace(variant, '')
                    element = element.rstrip()

                # Extract and remove "Engine" data
                engine_match = re.search(r'(\d+(\.\d+)?L .+)', element)
                if engine_match:
                    engine = engine_match.group(1)

                    # Assign the extracted "Engine" data to the respective column
                    engine_col = ws[f'Q{cell_row}']
                    engine_col.value = engine

                    # Remove "Engine" data from car description element
                    element = element.replace(engine, '')
                    element = element.rstrip()

                    # Split the remaining car description into parts
                    parts = element.split(" ")

                    # Assign each part to its respective column
                    make = parts[0]
                    model_series = " ".join(parts[1:])

                    # Assign the extracted parts data to their respective columns
                    make_col = ws[f'O{cell_row}']
                    mod_ser_col = ws[f'P{cell_row}']

                    make_col.value = make
                    mod_ser_col.value = model_series

        except:
            comment_col = ws[f'U{cell_row}']
            comment_col.value = f"Plate number not found."

        # Add timestamp to each Excel row
        timestamp = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        timestamp_col = ws[f'W{cell_row}']
        timestamp_col.value = timestamp

        wb.save(excel_file)
        driver.refresh()
        
driver.quit()

print("Web scraping complete.")
